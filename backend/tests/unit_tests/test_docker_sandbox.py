"""
Unit tests for DockerSandbox - Secure container execution environment

Tests cover:
1. Basic code execution (simple Python code)
2. File input/output (Excel file processing)
3. Security restrictions (network disabled, resource limits)
4. Error handling (syntax errors, runtime errors)
5. Container cleanup (no orphaned containers)
"""
import textwrap
import pytest
import io
from app.services.ai_executor.docker_sandbox import DockerSandbox, ExecutionResult
import docker


class TestDockerSandbox:
    """Test suite for DockerSandbox container execution"""

    def setup_method(self):
        """Initialize DockerSandbox before each test"""
        self.sandbox = DockerSandbox()

    def test_sandbox_initialization(self):
        """Test DockerSandbox initializes successfully"""
        assert self.sandbox.client is not None, "Docker client should be initialized"
        assert self.sandbox.IMAGE_NAME == "excel-ai-executor:latest"

        # Verify image exists
        try:
            image = self.sandbox.client.images.get(self.sandbox.IMAGE_NAME)
            assert image is not None, "Docker image should exist"
        except docker.errors.ImageNotFound:
            pytest.fail(f"Docker image '{self.sandbox.IMAGE_NAME}' not found. Run: docker build -f Dockerfile.executor -t excel-ai-executor:latest .")

    def test_simple_code_execution(self):
        """Test executing simple Python code (no files)"""
        code = """
print("Hello from Docker!")
print("2 + 2 =", 2 + 2)
"""

        result = self.sandbox.execute_code(code)

        assert result.success == True, "Simple code should execute successfully"
        assert result.exit_code == 0, "Exit code should be 0"
        assert "Hello from Docker!" in result.output, "Output should contain print statement"
        assert "2 + 2 = 4" in result.output, "Output should contain calculation result"
        assert result.error == "", "Should have no errors"

    def test_pandas_execution(self):
        """Test executing pandas code (verify library is installed)"""
        code = """
import pandas as pd
import numpy as np

# Create sample DataFrame
data = {'Name': ['Alice', 'Bob', 'Charlie'], 'Age': [25, 30, 35]}
df = pd.DataFrame(data)

print(df.to_string())
print(f"Mean age: {df['Age'].mean()}")
"""

        result = self.sandbox.execute_code(code)

        assert result.success == True, "Pandas code should execute"
        assert "Alice" in result.output, "Should contain DataFrame data"
        assert "Mean age: 30" in result.output, "Should calculate mean correctly"

    def test_file_input_output(self):
        """Test file input/output with Excel files"""
        # Create a simple CSV file (easier to test than Excel)
        csv_content = b"Name,Age\nAlice,25\nBob,30\n"

        code = """
import pandas as pd

# Read input file
df = pd.read_csv('/tmp/input/test.csv')

# Process data
df['Age_Plus_10'] = df['Age'] + 10

# Write output file
df.to_csv('/tmp/output/result.csv', index=False)

print(f"Processed {len(df)} rows")
"""

        result = self.sandbox.execute_code(
            code=code,
            input_files={"test.csv": csv_content}
        )

        assert result.success == True, "File processing should succeed"
        assert "Processed 2 rows" in result.output
        assert "result.csv" in result.output_files, "Should generate output file"

        # Verify output file content
        output_csv = result.output_files["result.csv"].decode('utf-8')
        assert "Age_Plus_10" in output_csv, "Output should have new column"
        assert "35" in output_csv, "Should calculate Alice's age + 10 correctly"

    def test_syntax_error_handling(self):
        """Test handling of code with syntax errors"""
        code = """
print("Missing closing parenthesis"
"""

        result = self.sandbox.execute_code(code)

        assert result.success == False, "Syntax error should fail"
        assert result.exit_code != 0, "Exit code should be non-zero"
        assert "SyntaxError" in result.error or "invalid syntax" in result.error

    def test_runtime_error_handling(self):
        """Test handling of runtime errors"""
        code = """
# This will cause a runtime error
result = 10 / 0
"""

        result = self.sandbox.execute_code(code)

        assert result.success == False, "Division by zero should fail"
        assert result.exit_code != 0
        assert "ZeroDivisionError" in result.error

    def test_import_error_handling(self):
        """Test handling of missing imports"""
        code = """
import nonexistent_library
"""

        result = self.sandbox.execute_code(code)

        assert result.success == False, "Missing import should fail"
        assert "ModuleNotFoundError" in result.error or "ImportError" in result.error

    def test_network_disabled(self):
        """Test that network access is disabled (security)"""
        code = """
import urllib.request

try:
    response = urllib.request.urlopen('https://google.com')
    print("Network access worked - SECURITY ISSUE!")
except Exception as e:
    print(f"Network blocked (expected): {type(e).__name__}")
"""

        result = self.sandbox.execute_code(code)

        # Should fail because network is disabled
        assert "Network blocked" in result.output or result.success == False

    def test_file_access_restrictions(self):
        """Test that file access is restricted to /tmp directories"""
        # Test that we can ONLY access /tmp/input and /tmp/output
        # This test verifies Docker-level file system isolation
        code =  textwrap.dedent("""
    from pathlib import Path

    # Test 1: Can we read from /tmp/input? (should work)
    input_path = Path('/tmp/input/test.txt')
    try:
        input_path.write_text('test data')
        content = input_path.read_text()
        print(f"✅ /tmp/input access: SUCCESS ({content})")
    except Exception as e:
        print(f"❌ /tmp/input access: FAILED - {type(e).__name__}")

    # Test 2: Can we write to /tmp/output? (should work)
    output_path = Path('/tmp/output/result.txt')
    try:
        output_path.write_text('output data')
        print("✅ /tmp/output access: SUCCESS")
    except Exception as e:
        print(f"❌ /tmp/output access: FAILED - {type(e).__name__}")

    # Test 3: Try to write outside /tmp (should fail - read-only filesystem outside /tmp)
        # Docker containers can be configured with read-only filesystems
        restricted_path = Path('/home/test_write.txt')
        try:
            restricted_path.write_text('malicious data')
            print("🚨 SECURITY ISSUE: Wrote outside /tmp!")
        except PermissionError:
            print("✅ Write blocked: PermissionError (expected)")
        except Exception as e:
            print(f"✅ Write blocked: {type(e).__name__}")
    """)
      
        result = self.sandbox.execute_code(code)
        
        # Should successfully access /tmp directories
        assert "✅ /tmp/input access: SUCCESS" in result.output, "Should access /tmp/input"
        assert "✅ /tmp/output access: SUCCESS" in result.output, "Should access /tmp/output"
        
        # Should NOT be able to write outside /tmp
        assert "SECURITY ISSUE" not in result.output, "Should not write outside /tmp"

    def test_container_cleanup(self):
        """Test that containers are properly cleaned up after execution"""
        # Get list of containers before execution
        containers_before = self.sandbox.client.containers.list(all=True)

        # Execute simple code
        code = "print('Testing cleanup')"
        result = self.sandbox.execute_code(code)

        assert result.success == True

        # Get list of containers after execution
        containers_after = self.sandbox.client.containers.list(all=True)

        # Should have same number of containers (our container was removed)
        assert len(containers_after) == len(containers_before), "Container should be cleaned up"

    def test_multiple_executions(self):
        """Test that multiple executions work independently (no state sharing)"""
        # First execution: Set variable
        code1 = """
test_var = "First execution"
print(test_var)
"""
        result1 = self.sandbox.execute_code(code1)
        assert result1.success == True
        assert "First execution" in result1.output

        # Second execution: Try to access variable from first (should fail - new container)
        code2 = """
try:
    print(test_var)
    print("SECURITY ISSUE: State persisted between executions!")
except NameError:
    print("Variable not found (expected - new container)")
"""
        result2 = self.sandbox.execute_code(code2)
        assert result2.success == True
        assert "Variable not found" in result2.output, "Containers should be isolated"

    def test_openpyxl_library(self):
        """Test that openpyxl library is available for Excel operations"""
        code = """
from openpyxl import Workbook

# Create workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
ws['B1'] = 'Excel'

# Save to output
wb.save('/tmp/output/test.xlsx')
print("Excel file created successfully")
"""

        result = self.sandbox.execute_code(code)

        assert result.success == True, "openpyxl should work"
        assert "Excel file created successfully" in result.output
        assert "test.xlsx" in result.output_files, "Should generate Excel file"

    def test_requests_library_available(self):
        """Test that requests library is installed (MEDIUM_RISK library)"""
        code = """
import requests

# Just verify import works (network is disabled, so can't make actual requests)
print(f"requests library version: {requests.__version__}")
print("requests library is available")
"""

        result = self.sandbox.execute_code(code)

        assert result.success == True, "requests should be installed"
        assert "requests library is available" in result.output
